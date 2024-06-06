#! /usr/bin/python
# Copyright (C) 2023 Robert Bosch GmbH
import os
import shutil
import glob
import math
import asyncio
import time
from prettytable import PrettyTable
import statistics
import csv
import argparse
import sys
from pathlib import Path
import yaml
try: # below packages are not needed/available in 'dse-python-builder' container image
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import SurfaceChart3D, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    from dse.opstools.schemas.signalgroup import SignalGroup, Signal
    from dse.opstools.schemas.model import Model, Channel, RuntimeSpec
    from dse.opstools.schemas.stack import Stack, ModelInstance, ModelInstanceSpec
    from dse.opstools.schemas.types import GenericDict
except Exception as e:
    print(e)

path = Path(os.getcwd())
BASE_PATH = str(path.parent.parent.parent.absolute())
BENCHMARK_OUT_PATH = BASE_PATH + '/dse/modelc/build/_out/examples/benchmark'

TIMEOUT = 6000
VALGRIND_CMD = 'valgrind --track-origins=yes --leak-check=full --error-exitcode=808'
USE_VALGRIND = False
CWD = os.getcwd() + "/"

# Sandbox for Model C
MODELC_SANDBOX_DIR = CWD + 'dse/modelc/build/_out'
SIMBUS_EXE = MODELC_SANDBOX_DIR+'/bin/simbus'
MODELC_EXE = MODELC_SANDBOX_DIR+'/bin/modelc'
WORKING_DIR = CWD + 'tests/pytest/benchmark/_working/'
MODEL_YAML = WORKING_DIR + 'model.yaml'
BENCHMARK_RESULT = {}
TCP_TRANSPORT = 'redis://redis:6379'
SOCKET_TRANSPORT = 'unix:///tmp/redis/redis.sock'
SIMER_SOCKET_TRANSPORT = 'unix:///var/run/redis/redis.sock'
MQ_TRANSPORT = 'posix:///stem'
STACK_YAML = WORKING_DIR + "stack.yaml" # Default stack yaml
SIGNAL_GROUP_YAML = WORKING_DIR + "signal_group.yaml" # Default signal_group yaml
USE_SAME_MODEL_SOURCE_OBJECT = False # Update value to 'True' if each model instance in stack.yaml file to point same model source object.
CHANNEL_NAMES = os.getenv('CHANNEL_NAMES', 'data_1_channel;')
CHANNEL_ALIAS = os.getenv('CHANNEL_NAMES', 'data_1;')
SIMER_SIMULATION_PATH = '_working/simer'
LOG_LEVEL = 1


def generate_signal_group_yaml(num_of_signals):
    start_index = 0
    signal_limit = num_of_signals = int(num_of_signals)
    signalgroup_path = "_working/signal_group.yaml"
    open(signalgroup_path, 'w').close() # removes the file content
    for chnl_name in CHANNEL_NAMES.split(';'):
        signals = ['signal_' + str(i) for i in range(start_index, signal_limit)]
        signal_group = SignalGroup()
        signal_group.metadata.name = chnl_name
        signal_group.metadata.labels['channel'] = chnl_name
        for signal in signals:
            signal_group.spec.signals.append(Signal(signal))
        with open(signalgroup_path, 'a') as md:
            md.write(signal_group.dump())
        start_index = signal_limit
        signal_limit = signal_limit + num_of_signals


def read_input_file():
    with open('input.csv', 'r') as file:
        input_data = file.readlines()[1:]
    return input_data


def generate_channel_names(num_of_channels):
    global CHANNEL_NAMES, CHANNEL_ALIAS
    CHANNEL_NAMES = CHANNEL_ALIAS = ''
    for i in range(1, int(num_of_channels) + 1):
        CHANNEL_NAMES += 'data_' + str(i) + '_channel;'
        CHANNEL_ALIAS += 'data_' + str(i) + ';'
    CHANNEL_NAMES = CHANNEL_NAMES.rstrip(';')
    CHANNEL_ALIAS = CHANNEL_ALIAS.rstrip(';')
    return CHANNEL_NAMES, CHANNEL_ALIAS


def generate_model_yaml(num_of_channels):
    model_path = "_working/model.yaml"
    open(model_path, 'w').close() # removes the file content
    model = Model()
    model.metadata.name = 'Benchmark'
    channel_names, channel_alias = generate_channel_names(num_of_channels)
    for c_name, c_alias in zip(channel_names.split(';'), channel_alias.split(';')):
        channel = Channel(alias=c_alias, selectors=GenericDict({'channel': c_name}))
        model.spec.channels.append(channel)
    model_name = 'libbenchmark'
    model.spec.runtime.dynlib.append(RuntimeSpec(path=f'lib/{model_name}.so', os='linux', arch='amd64'))
    model.spec.runtime.dynlib.append(RuntimeSpec(path=f'lib/{model_name}.so', os='linux', arch='x86'))
    model.spec.runtime.dynlib.append(RuntimeSpec(path=f'bin/{model_name}.dll', os='windows', arch='x64'))
    model.spec.runtime.dynlib.append(RuntimeSpec(path=f'bin/{model_name}.dll', os='windows', arch='x86'))
    with open(model_path, 'a') as md:
        md.write(model.dump())


def create_simer_simulation_dir():
    os.makedirs('_working/simer/data', exist_ok=True)
    os.makedirs('_working/simer/lib', exist_ok=True)


def setup_simer_simulation_files():
    shutil.copy('_working/model.yaml', '_working/simer/data')
    shutil.copy(BENCHMARK_OUT_PATH + '/lib/libbenchmark.so', '_working/simer/lib')
    shutil.copy('_working/simulation.yaml', '_working/simer/data')
    shutil.copy('_working/signal_group.yaml', '_working/simer/data')


def update_yaml(path):
    updated_docs = []
    with open(path, 'r') as fileStr:
        docs = yaml.safe_load_all(fileStr)
        for doc in docs:
            if doc['kind'] == 'Stack':
                if TOPOLOGY == 'stacked':
                    doc['spec'].update({'runtime': {'stacked': True}})
                if TRANSPORT in (SOCKET_TRANSPORT, TCP_TRANSPORT):
                    doc['spec']['connection']['transport']['redispubsub'].update({'timeout': 60, 'uri': 'redis://localhost:6379'})
                elif TRANSPORT == SIMER_SOCKET_TRANSPORT:
                    doc['spec']['connection']['transport']['redispubsub'].update({'timeout': 60, 'uri': SIMER_SOCKET_TRANSPORT})
                for i, model in enumerate(doc['spec']['models'], start = 0):
                    if model['name'] != 'simbus':
                        doc['spec']['models'][i].update({'runtime': {'files': ['data/signal_group.yaml']}})
            updated_docs.append(doc)
    with open('_working/simulation.yaml', 'w') as md:
        yaml.safe_dump_all(updated_docs, md)


def generate_stack_yaml(num_of_models, TRANSPORT, num_of_channels):
    ''' function to update the stack.yaml file with dynamically generated model instance names data'''
    stack = Stack()
    stack.metadata['name'] = 'benchmark_stack'
    uri = 'redis://redis:6379'
    if TRANSPORT == MQ_TRANSPORT:
        uri = MQ_TRANSPORT
    elif TRANSPORT == SOCKET_TRANSPORT:
        uri = SOCKET_TRANSPORT
    elif TRANSPORT == SIMER_SOCKET_TRANSPORT:
        uri = SIMER_SOCKET_TRANSPORT

    channel_names, channel_alias = generate_channel_names(num_of_channels)
    channel_names = channel_names.split(';')
    channel_alias = channel_alias.split(';')
    stack.set_redispubsub(uri=uri, timeout=60)

    # if RUNTIME == 'simer':
    #     stack.spec.runtime['stacked'] = True
    #     # stack.spec.update({'runtime':{'stacked':True}})

    stack.spec.models.append(
                ModelInstance(
                    name='simbus',
                    model = ModelInstanceSpec(name='simbus'),
                    channels = [Channel(name=chnl, expectedModelCount=num_of_models) for chnl in channel_names],
                )
            )
    uid = 43
    for i in range(1, int(num_of_models) + 1):
        stack.spec.models.append(
                    ModelInstance(
                        name= 'benchmark_inst_' + str(i),
                        uid=uid,
                        model = ModelInstanceSpec(name='Benchmark'),
                        channels = [Channel(name=chnl, alias=chnl_alias) for (chnl, chnl_alias) in zip(channel_names, channel_alias)],
                    ))
        uid += 1

    model = Model()
    model.metadata.name = 'simbus'
    with open('_working/stack.yaml', 'w') as md:
        stack_data = Stack.load(stack.dump()).dump()
        if TRANSPORT == MQ_TRANSPORT:
            stack_data = stack_data.replace('redispubsub', 'mq')
        md.write(stack_data)
        md.write(model.dump())
    if RUNTIME == 'simer': # to add the 'runtime' key and val in stack spec
        update_yaml('_working/stack.yaml')


def raise_exception(msg):
    raise Exception(msg)


def check_type(value, check_if):
    status = False
    if check_if == 'int':
        try :
            int(value)
        except ValueError:
            status = True
    if check_if == 'float':
        try :
            float(value)
        except ValueError:
            status = True
    return status


def validate_input(simbus_uri, num_of_models, step_size, end_time, signal_count, signal_change):
    simbus_uri = simbus_uri.strip()
    if simbus_uri not in ('redis://redis:6379', 'unix:///tmp/redis/redis.sock', 'posix:///stem', 'unix:///var/run/redis/redis.sock') :
        raise_exception('Invalid simbus URI.')
        return False
    if check_type(num_of_models, 'int'):
        raise_exception('num_of_models value should be integer')
        return False
    if check_type(step_size, 'float'):
        raise_exception('step_size value should be float/int')
        return False
    if check_type(end_time, 'float'):
        raise_exception('end_time value should be float/int')
        return False
    if check_type(signal_count, 'int'):
        raise_exception('signal_count value should be integer')
        return False
    if check_type(signal_change, 'int'):
        raise_exception('signal_change value should be integer')
        return False


def run_benchmark(input_data):
    global TRANSPORT, NUM_OF_CHANNELS, SIGNAL_COUNT, SIGNAL_CHANGE, NUM_OF_MODELS, STEP_SIZE, END_TIME
    prev_signal_count = 0
    prev_model_count = 0
    for line in input_data:
        data = line.split(',')
        if '' in data :
            continue # checks if any col value missing or any empty row in inpuit csv file

        if '-CMD_IN' not in TRANSPORT:
            TRANSPORT = data[0]
        if '-CMD_IN' not in str(NUM_OF_CHANNELS):
            NUM_OF_CHANNELS = data[1]
        if '-CMD_IN' not in str(NUM_OF_MODELS):
            NUM_OF_MODELS = data[2]
        if '-CMD_IN' not in str(STEP_SIZE):
            STEP_SIZE = data[3]
        if '-CMD_IN' not in str(END_TIME):
            END_TIME = data[4]
        if '-CMD_IN' not in str(SIGNAL_COUNT):
            SIGNAL_COUNT = data[5]
        if '-CMD_IN' not in str(SIGNAL_CHANGE):
            SIGNAL_CHANGE = data[6]
        OUT_FILE = data[7]

        TRANSPORT = TRANSPORT.replace('-CMD_IN', '')
        NUM_OF_CHANNELS = NUM_OF_CHANNELS.replace('-CMD_IN', '')
        SIGNAL_COUNT = SIGNAL_COUNT.replace('-CMD_IN', '')
        SIGNAL_CHANGE = SIGNAL_CHANGE.replace('-CMD_IN', '')
        NUM_OF_MODELS = NUM_OF_MODELS.replace('-CMD_IN', '')
        STEP_SIZE = STEP_SIZE.replace('-CMD_IN', '')
        END_TIME = END_TIME.replace('-CMD_IN', '')

        if validate_input(TRANSPORT, NUM_OF_MODELS, STEP_SIZE, END_TIME, SIGNAL_COUNT, SIGNAL_CHANGE) is False :
            continue

        generate_stack_yaml(NUM_OF_MODELS, TRANSPORT, NUM_OF_CHANNELS)
        generate_model_yaml(NUM_OF_CHANNELS)

        # creates .env file to set the env variables for docker-compose.yaml
        with open('_working/.env', 'w') as file:
            content = [
                f'TRANSPORT={TRANSPORT}\n',
                f'NUM_OF_MODELS={NUM_OF_MODELS}\n',
                f'STEP_SIZE={STEP_SIZE}\n',
                f'END_TIME={END_TIME}\n',
                f'SIGNAL_COUNT={SIGNAL_COUNT}\n',
                f'TOPOLOGY={TOPOLOGY}\n',
                f'SIGNAL_CHANGE={SIGNAL_CHANGE}\n',
                f'NUM_OF_CHANNELS={NUM_OF_CHANNELS}\n',
                f'RUNTIME={RUNTIME}\n',
                f'OUTDIR={OUTDIR}\n',
                f'OUT_FILE={OUT_FILE}']
            file.writelines(content)

        # generates signal_group.yaml file only if there is a change in signal_count from the above row in input.csv
        if SIGNAL_CHANGE != prev_signal_count or NUM_OF_MODELS != prev_model_count:
            generate_signal_group_yaml(SIGNAL_CHANGE)
            prev_signal_count = SIGNAL_CHANGE
            prev_model_count = NUM_OF_MODELS
        if RUNTIME == 'simer':
            create_simer_simulation_dir()
            setup_simer_simulation_files()
            with open('_working/.env') as env_lines:
                for line in env_lines:
                    env_keyval = line.split('=')
                    os.environ[env_keyval[0]] = env_keyval[1].replace('\n', '')
            os.system('make simer')
        else:
            os.system('make benchmark')


def process_data_for_plotting(df):
    signal_change_range = df['signal change'].unique()
    gp_df = df.groupby(['scenario', 'transport', 'channels', 'models'])
    cols = ['scenario', 'transport', 'channels', '']
    cols.extend([''] * len(signal_change_range))
    data_list = []
    prev_transport_name = ''
    prev_scenario_name = ''
    prev_channel_count = ''
    for gp_by, group in gp_df:
        row_items = [gp_by[0], gp_by[1], gp_by[2], gp_by[3]]
        for row in group.iterrows():
            sim_steps_per_sec = row[1]['rate (steps/sec)']
            row_items.extend([sim_steps_per_sec])
        if prev_transport_name != gp_by[1] or prev_scenario_name != gp_by[0] or prev_channel_count != gp_by[2]:
            tmp_header = ['', '', '', '']
            tmp_header.extend(signal_change_range)
            data_list.append(cols)
            data_list.append(tmp_header)
            prev_scenario_name = gp_by[0]
            prev_transport_name = gp_by[1]
            prev_channel_count = gp_by[2]
        data_list.append(row_items)
    data_list.append(data_list[0]) # needed for splitting to different dfs
    data_list = data_list[1:]
    df_list = []
    tmp_list = []
    for row_list in data_list:
        if row_list[0] == 'scenario':
            tmp_df = pd.DataFrame(tmp_list, columns = cols)
            df_list.append(tmp_df)
            tmp_list.clear()
        else:
            tmp_list.append(row_list)
    return df_list


def plot_to_excel(plot_data, csv_data, max_df_val, file_name):
    scenario = plot_data['scenario'][1]
    transport = plot_data['transport'][1]
    channels = plot_data['channels'][1]
    plot_data = plot_data.drop(columns = ['scenario', 'transport', 'channels'])
    row_count, col_count = plot_data.shape
    out_file_name = f'_out/{file_name}.xlsx'

    tmp_scenario = scenario.split('_')
    sheet_name = (tmp_scenario[-1].title() + tmp_scenario[1].title() + transport.replace('redispubsub_', 'redis_').replace('message_queue', 'mq').title() + '#' + str(1)).replace('_', '')

    if os.path.exists(out_file_name) is False :
        wb = Workbook()
        ws = wb.active
        ws.title = f'{sheet_name}_D' # gets created only one time
        wb.save(out_file_name)
    wb = load_workbook(out_file_name)
    sheetNames = [name.replace('_P', '') for name in wb.sheetnames if '_D' not in name]
    if len(sheetNames) > 0:
        sheetNameId = sheetNames[-1].split('#')
        sheet_id = int(sheetNameId[-1]) + 1
        sheet_name = sheetNameId[0] + '#' + str(sheet_id)

    data_sheet = sheet_name + '_D'
    if data_sheet not in wb.sheetnames:
        wb.create_sheet(data_sheet)
    ws = wb[data_sheet]
    for row in dataframe_to_rows(csv_data, index=False, header=True):# loading the complete plot data
        ws.append(row)

    plot_sheet = sheet_name + '_P'
    wb.create_sheet(plot_sheet)
    ws = wb[plot_sheet]
    data = plot_data.values.tolist()
    for row in data:
        ws.append(row)
    chart = SurfaceChart3D()
    labels = Reference(ws, min_col = 1, min_row = 2, max_row = row_count)
    data = Reference(ws, min_col = 2, max_col = col_count, min_row = 1, max_row = row_count)
    chart.add_data(data, titles_from_data = True)
    chart.set_categories(labels)
    chart.title = (scenario.replace('_', ' ') + ' - ' + transport.replace('_', ' ') + ',channels-'+str(channels)).title()
    chart.style = 26
    chart.height = 10 # default is 7.5
    chart.width = 20 # default is 15
    chart.y_axis.title = "Sim. Steps per sec."
    chart.x_axis.title = "Num. Of Models"
    chart.z_axis.title = "Signal Change"

    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = max_df_val
    # chart.y_axis.scaling.unit = 100
    if os.path.exists(f"_out/{file_name}.csv"):
        os.remove(f"_out/{file_name}.csv")
    ws.add_chart(chart, "K2")
    wb.active = wb[plot_sheet]
    wb.save(out_file_name)


def benchmark():
    input_data = read_input_file()
    run_benchmark(input_data)
    print('Plotting Surface charts...')
    out_files = [file for file in glob.glob("_out/*.csv")] # input files

    def roundup(x):
        '''rounding the max y-axis value to its closest hundreds'''
        return int(math.ceil(x / 100.0)) * 100

    def get_max_df_value(df_list):
        '''To fetch the max value from dataframe to sync y-axis values across plots'''
        max_df_val = 0
        for df in df_list:
            df_max = df.iloc[1:].drop(df.columns[[0, 1]], axis=1).to_numpy().max()
            max_df_val = df_max if df_max > max_df_val else max_df_val
        max_df_val = roundup(max_df_val)
        return max_df_val

    for file in out_files:
        file_name = file.split('/')[-1].rstrip('.csv')
        csv_data = pd.read_csv(file)
        df_list = process_data_for_plotting(csv_data)
        max_df_val = get_max_df_value(df_list)
        for df in df_list:
            plot_to_excel(df, csv_data, max_df_val, file_name)
    print(f'Plotted Surface charts.')


def move_output_file():
    if OUTDIR != default_outdir :
        in_data = pd.read_csv(SCENARIO)
        in_filenames = [x.replace('.csv', '.xlsx') for x in in_data['out_file'].unique().tolist()]
        file_names = os.listdir(default_outdir)
        for file_name in file_names:
            if file_name in in_filenames:
                shutil.move(os.path.join(default_outdir, file_name), OUTDIR.rstrip('/').rstrip('\n') + '/' + file_name)


def copy_input_file():
    if hasattr(args, 'scenario') and args.scenario is not None:
        try:
            shutil.copy(SCENARIO.replace('\n', ''), CWD + 'input.csv')
        except shutil.SameFileError:
            pass


def find_path(path):
    if not os.path.exists(path):
        os.mkdir(path)
    if path[-1] != '/':
        path += '/'
    return path


def form_model_instance_param_data(NUM_OF_MODELS):
    INDIVIDUAL_MODEL_INSTANCE_MODELS, STACKED_MODEL_INSTANCE_MODEL_INSTANCES = [], ''
    for i in range(1, int(NUM_OF_MODELS) + 1):
        model_inst = 'benchmark_inst_' + str(i)
        instance_dict = {
                    'MODEL_INST': model_inst,
                    'MODEL_YAML_FILES': f'{MODEL_YAML} ' + f'{STACK_YAML} ' + f'{SIGNAL_GROUP_YAML}',
                }
        INDIVIDUAL_MODEL_INSTANCE_MODELS.append(instance_dict)
        STACKED_MODEL_INSTANCE_MODEL_INSTANCES += model_inst + ';'
    STACKED_MODEL_INSTANCE_MODEL_INSTANCES = '"' + STACKED_MODEL_INSTANCE_MODEL_INSTANCES.rstrip(';') + '"'
    return INDIVIDUAL_MODEL_INSTANCE_MODELS, STACKED_MODEL_INSTANCE_MODEL_INSTANCES


async def run(dir, cmd):
    _cmd = f'cd {dir}; {VALGRIND_CMD if USE_VALGRIND else ""} {cmd}'
    print('Run:')
    print(f'  dir: {dir}')
    print(f'  cmd: {_cmd}')
    p = await asyncio.create_subprocess_shell(
        _cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await p.communicate()
    return { 'dir': dir, 'cmd': _cmd, 'rc': p.returncode, 'stdout': stdout.decode(), 'stderr': stderr.decode() }


async def main(params, checks):
    run_list = []
    uri = TRANSPORT

    transport_name = 'redispubsub'
    transport_uri_str = ''
    if uri == SOCKET_TRANSPORT:
        transport_uri_str = f'--uri {uri} '
    elif uri == MQ_TRANSPORT:
        transport_uri_str = f'--uri {uri} '
        transport_name = 'mq'

    # SimBus
    run_list.append(asyncio.wait_for(run(
            params['MODEL_SANDBOX_DIR'], f'{SIMBUS_EXE} --logger {LOG_LEVEL} ' + f'--stepsize {STEP_SIZE} ' + f'--transport {transport_name} ' + transport_uri_str + f'--timeout 1 {STACK_YAML}'), timeout=TIMEOUT)
    )
    # Models
    for m in params['models']:
        run_list.append(asyncio.wait_for(run(
            params['MODEL_SANDBOX_DIR'], f'{MODELC_EXE} --logger {LOG_LEVEL} ' + f'--name {m["MODEL_INST"]} ' + f'{m["MODEL_YAML_FILES"]} ' + f'--stepsize {STEP_SIZE} ' + f'--transport {transport_name} ' + transport_uri_str + f'--endtime {END_TIME}'), timeout=TIMEOUT))
    # Gather results.
    result_list = await asyncio.gather(*run_list)

    for result in result_list:
        print('************************************************************')
        print('************************************************************')
        print('************************************************************')
        print(f'CMD : {result["cmd"]}')
        print('++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
        print(f'STDOUT : \n{result["stdout"]}')
        print('++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
        print(f'STDERR : \n{result["stderr"]}')
        print('++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
        print(f'RC : {result["rc"]}')
        print('************************************************************')
        print('')
        assert result['rc'] == 0

    for check in checks:
        # Search for check in each result STDOUT
        found = False
        for result in result_list:
            if check in result['stdout']:
                found = True
        assert found, check


def exec_time(func_name, key, value):
    if func_name in BENCHMARK_RESULT:
        dct_obj = BENCHMARK_RESULT[func_name]
        dct_obj[key].append(value)
    else :
        BENCHMARK_RESULT[func_name] = { key : [value] } # time_in_sec


def generate_csv(csv_file_data, out_path):
    if os.path.isdir(out_path) is False :
        os.mkdir(out_path)
    with open(out_path + '/' + OUT_FILE, 'a', newline='') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        for row in csv_file_data:
            csvwriter.writerow(row)


def display_benchmark_result():
    if RUNTIME == 'docker':
        out_path = CWD + "tests/pytest/benchmark/_out"
    else:
        out_path = CWD + "_out"
    log_table = PrettyTable()
    print('================================================ Benchmark Test Result ================================================')
    field_names = ["runtime", "scenario", "transport", "channels", "models", "step size", "end time", "signals", "signal change" , "run time", "rate (steps/sec)", "factor ()"]
    log_table.field_names = field_names
    csv_file_data = [] if os.path.isfile(out_path + '/' + OUT_FILE) else [field_names] # adding csv header data
    for func in BENCHMARK_RESULT:
        exec_time_list = BENCHMARK_RESULT[func]['runtime']
        # min_runtime = "{:.6f}".format(min(exec_time_list))
        mean_runtime = "{:.6f}".format(statistics.mean(exec_time_list))
        # median_runtime = "{:.6f}".format(statistics.median(exec_time_list))
        # max_runtime = "{:.6f}".format(max(exec_time_list))
        sim_step_per_second = "{:.0f}".format((float(END_TIME) / float(STEP_SIZE)) / float(mean_runtime))
        speed = "{:.4f}".format(float(END_TIME) / float(mean_runtime))
        transport = 'redispubsub_tcp'
        if TRANSPORT == MQ_TRANSPORT:
            transport = 'message_queue'
        elif TRANSPORT == SOCKET_TRANSPORT or TRANSPORT == SIMER_SOCKET_TRANSPORT:
            transport = 'redispubsub_socket'
        row = [RUNTIME, func, transport, NUM_OF_CHANNELS, NUM_OF_MODELS, STEP_SIZE, END_TIME, SIGNAL_COUNT , SIGNAL_CHANGE, mean_runtime, sim_step_per_second, speed]
        csv_file_data.append(row)
        log_table.add_row(row)
    generate_csv(csv_file_data, out_path)
    print(log_table)


def timer_func(func):
    def function_timer():
        start = time.time()
        value = func()
        end = time.time()
        runtime = end - start
        exec_time(func.__name__, 'runtime', runtime)
        return value
    return function_timer


@timer_func
def test_individual_model_instance():
    # Two models executed by the two instances of ModelC.
    params = {
        'MODEL_SANDBOX_DIR': MODELC_SANDBOX_DIR + "/examples/benchmark",
        'models': INDIVIDUAL_MODEL_INSTANCE_MODELS
    }
    checks = []
    asyncio.run(main(params, checks))


@timer_func
def test_stacked_model_instance():
    # Two models executed by the same instance of ModelC.
    params = {
        'MODEL_SANDBOX_DIR': MODELC_SANDBOX_DIR + '/examples/benchmark',
        'models': [
            {
                'MODEL_INST': STACKED_MODEL_INSTANCE_MODEL_INSTANCES,
                'MODEL_YAML_FILES': f'{MODEL_YAML} ' + f'{STACK_YAML} ' + f'{SIGNAL_GROUP_YAML}',
            },
        ]
    }
    checks = []
    asyncio.run(main(params, checks))


@timer_func
def test_individual_model_instance_simer():
    os.system(SIMER_SHELL_CMDS)


@timer_func
def test_stacked_model_instance_simer():
    os.system(SIMER_SHELL_CMDS)


def create_simer_shell_cmds():
    global SIMER_SHELL_CMDS
    transport_name = 'redispubsub'
    transport_uri_str = 'redis://localhost:6379'
    if TRANSPORT == SOCKET_TRANSPORT:
        transport_uri_str = TRANSPORT
    elif TRANSPORT == SIMER_SOCKET_TRANSPORT:
        transport_uri_str = TRANSPORT
    elif TRANSPORT == MQ_TRANSPORT:
        transport_uri_str = TRANSPORT
        transport_name = 'mq'
    SIMER_SHELL_CMDS = f'''export SIMER_IMAGE=ghcr.io/boschglobal/dse-simer:latest;
                        simer() {{ ( cd "$1" && shift && docker run -it --rm -v $(pwd):/sim $SIMER_IMAGE "$@"; ) }};
                        simer {SIMER_SIMULATION_PATH} -endtime {END_TIME} -stepsize {STEP_SIZE} -transport {transport_name} -uri {transport_uri_str} -logger {LOG_LEVEL}'''
    print(SIMER_SHELL_CMDS.split(';')[-1].strip())


def execute_tests():
    if args.dockerexec:
        if TOPOLOGY == 'stacked':
            test_stacked_model_instance()
        else:
            test_individual_model_instance()
    elif args.simerexec:
        create_simer_shell_cmds()
        if TOPOLOGY == 'stacked':
            test_stacked_model_instance_simer()
        else:
            test_individual_model_instance_simer()
    display_benchmark_result()


default_transport = 'redis://redis:6379'
default_scenario = CWD + 'input.csv'
default_topology = 'individual'
default_outdir = CWD + '_out'
default_stepsize = '0.005'
default_endtime = '10'
default_outfile = 'benchmark_result.csv'
default_num_of_models = '1'
default_signal_count = 10
default_signal_change = default_signal_count
default_num_of_channels = 1
default_runtime = 'docker'

parser = argparse.ArgumentParser()

sub_parsers = parser.add_subparsers(dest="command")
run_parser = sub_parsers.add_parser("run", help="benchmark function execution with default argument values")

# parser for command 'generate'
generate_parser = sub_parsers.add_parser("generate", help="generate static yaml files in _workdir directory")
generate_parser.add_argument("--transport", nargs='?', choices=['redis://redis:6379', 'unix:///tmp/redis/redis.sock', 'posix:///stem'], help="Transport type")
generate_parser.add_argument("--signals", nargs='?', help="number of signals")
generate_parser.add_argument("--channels", nargs='?', help="number of channels")
generate_parser.add_argument("--models", nargs='?', help="number of models")
generate_parser.add_argument("--outdir", nargs='?', help="folder to save generated yaml files")

# parser for command 'execute'
execute_parser = sub_parsers.add_parser("execute", help="benchmark function execution which takes optional arguments to change transport, scenario, topology, outdir, signals, signal_change, channel, models, step size and end time.")
execute_parser.add_argument("--transport", nargs='?', choices=['redis://redis:6379', 'unix:///tmp/redis/redis.sock', 'posix:///stem', 'unix:///var/run/redis/redis.sock'], help="Transport type")
execute_parser.add_argument("--scenario", nargs='?',  help="input file path, file type : csv")
execute_parser.add_argument("--topology", nargs='?', choices=['individual', 'stacked'], help="")
execute_parser.add_argument("--outdir", nargs='?', help="folder to save generated result files")
execute_parser.add_argument("--signals", nargs='?', help="number of signals")
execute_parser.add_argument("--signal_change", nargs='?', help="number of signal change in each step")
execute_parser.add_argument("--channels", nargs='?', help="number of channels")
execute_parser.add_argument("--models", nargs='?', help="number of models")
execute_parser.add_argument("--step_size", nargs='?', help="step size")
execute_parser.add_argument("--end_time", nargs='?', help="end time")
execute_parser.add_argument("--runtime", nargs='?', help="runtime(simer or docker)")

run_parser.add_argument("--dockerexec", nargs='?', help=argparse.SUPPRESS, default=False)
run_parser.add_argument("--simerexec", nargs='?', help=argparse.SUPPRESS, default=False)
args = parser.parse_args()

# display help if no command given
if len(sys.argv) == 1:
    parser.print_help(sys.stderr)
    sys.exit(1)

STEP_SIZE = default_stepsize
END_TIME = default_endtime
OUT_FILE = default_outfile
NUM_OF_MODELS = default_num_of_models
SIGNAL_COUNT = default_signal_count
SIGNAL_CHANGE = default_signal_change
NUM_OF_CHANNELS = default_num_of_channels

TRANSPORT = default_transport
SCENARIO = default_scenario
TOPOLOGY = default_topology
OUTDIR = default_outdir
RUNTIME = os.getenv('RUNTIME', default_runtime)
if RUNTIME == 'simer':
    sh_transport = os.getenv('TRANSPORT')
    if sh_transport == 'redispubsub_tcp':
        TRANSPORT = 'redis://redis:6379' + '-CMD_IN'
        os.environ['TRANSPORT'] = TRANSPORT
    elif sh_transport == 'mq':
        TRANSPORT = 'posix:///stem' + '-CMD_IN'
        os.environ['TRANSPORT'] = TRANSPORT
    elif sh_transport == 'redispubsub_socket':
        TRANSPORT = 'unix:///var/run/redis/redis.sock' + '-CMD_IN'
        os.environ['TRANSPORT'] = TRANSPORT


def exec():
    try:
        copy_input_file()
        benchmark()
        move_output_file()
    except Exception as e:
        print('ERROR : ', e)


if args.command == "run":
    if args.dockerexec is False and args.simerexec is False:
        exec()
    else:
        NUM_OF_MODELS = os.getenv('NUM_OF_MODELS')
        STEP_SIZE = os.getenv('STEP_SIZE')
        END_TIME = os.getenv('END_TIME')
        SIGNAL_COUNT = os.getenv('SIGNAL_COUNT')
        SIGNAL_CHANGE = os.getenv('SIGNAL_CHANGE')
        OUT_FILE = os.getenv('OUT_FILE')
        TRANSPORT = os.getenv('TRANSPORT')
        TOPOLOGY = os.getenv('TOPOLOGY')
        OUTDIR = os.getenv('OUTDIR')
        NUM_OF_CHANNELS = os.getenv('NUM_OF_CHANNELS')
        if args.dockerexec:
            INDIVIDUAL_MODEL_INSTANCE_MODELS, STACKED_MODEL_INSTANCE_MODEL_INSTANCES = form_model_instance_param_data(NUM_OF_MODELS)
        execute_tests()

if args.command == "execute":
    os.makedirs('_working', exist_ok=True)
    if hasattr(args, 'transport') and args.transport is not None:
        TRANSPORT = args.transport + '-CMD_IN'
    if hasattr(args, 'scenario') and args.scenario is not None:
        SCENARIO = args.scenario
        SCENARIO = find_path(SCENARIO).rstrip('\n')
    if hasattr(args, 'topology') and args.topology is not None:
        TOPOLOGY = args.topology
    if hasattr(args, 'outdir') and args.outdir is not None:
        OUTDIR = args.outdir
        OUTDIR = find_path(OUTDIR)
    if hasattr(args, 'signals') and args.signals is not None:
        SIGNAL_COUNT = args.signals + '-CMD_IN'
    if hasattr(args, 'signal_change') and args.signal_change is not None:
        SIGNAL_CHANGE = args.signal_change + '-CMD_IN'
    if hasattr(args, 'channels') and args.channels is not None:
        NUM_OF_CHANNELS = args.channels + '-CMD_IN'
    if hasattr(args, 'models') and args.models is not None:
        NUM_OF_MODELS = args.models + '-CMD_IN'
    if hasattr(args, 'step_size') and args.step_size is not None:
        STEP_SIZE = args.step_size + '-CMD_IN'
    if hasattr(args, 'end_time') and args.end_time is not None:
        END_TIME = args.end_time + '-CMD_IN'
    if hasattr(args, 'runtime') and args.runtime is not None:
        RUNTIME = args.runtime
    if RUNTIME == 'simer':
        models = NUM_OF_MODELS
        signals = SIGNAL_CHANGE
        channels = NUM_OF_CHANNELS
        # variable type will be string if its from cmdline arguments, default is of type int
        if isinstance(models, str):
            models = models.rstrip('-CMD_IN')
        if isinstance(signals, str):
            signals = signals.rstrip('-CMD_IN')
        if isinstance(channels, str):
            channels = channels.rstrip('-CMD_IN')
        generate_stack_yaml(models, TRANSPORT.rstrip('-CMD_IN'), signals)
        generate_signal_group_yaml(signals)
        generate_model_yaml(channels)
    exec()

if args.command == "generate":
    dir_name = "_working"
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)
    channel = signal = model = 1 # default values
    transport = TCP_TRANSPORT # default value

    if hasattr(args, 'transport') and args.transport is not None:
        transport = args.transport
    if hasattr(args, 'signals') and args.signals is not None:
        signal = args.signals
    if hasattr(args, 'channels') and args.channels is not None:
        channel = args.channels
    if hasattr(args, 'models') and args.models is not None:
        model = args.models
    if hasattr(args, 'outdir') and args.outdir is not None:
        outdir = args.outdir
        outdir = find_path(outdir).replace('\n', '')
    else:
        outdir = CWD

    generate_stack_yaml(model, transport, channel)
    generate_signal_group_yaml(signal)
    generate_model_yaml(channel)

    try :
        shutil.copy('_working/stack.yaml', outdir + 'stack.yaml')
    except shutil.SameFileError:
        pass
    try :
        shutil.copy('_working/model.yaml', outdir + 'model.yaml')
    except shutil.SameFileError:
        pass
    try :
        shutil.copy('_working/signal_group.yaml', outdir + 'signal_group.yaml')
    except shutil.SameFileError:
        pass
    print('yaml files are generated.')
