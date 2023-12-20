import os
import signal_gen
import pandas as pd
import plotly.graph_objects as go
import glob
import math
from openpyxl import Workbook, load_workbook
from openpyxl.chart import SurfaceChart3D, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from ruamel import yaml
yaml = yaml.YAML(typ='rt')
path = Path(os.getcwd())
cmake_file_path = str(path.parent.parent.parent.absolute()) + '/dse/modelc/examples/benchmark/CMakeLists.txt'

USE_SAME_MODEL_SOURCE_OBJECT = False # Update value to 'True' if each model instance in stack.yaml file to point same model source object.


def read_input_file():
    with open('input.csv', 'r') as file:
        input_data = file.readlines()[1:]
    return input_data


def generate_model_instance_names(num_of_models):
    ''' Generate model instance names dynamically based on signal count.'''
    MODEL_INSTANCE_NAMES = ''
    for i in range(1, int(num_of_models) + 1):
        MODEL_INSTANCE_NAMES += 'model_instance_' + str(i) + ';'
    return MODEL_INSTANCE_NAMES.rstrip(';')


def update_cmakelist(model_instance_names):
    ''' Updates CMakeLists.txt file dynamically with new model instance data.'''
    with open(cmake_file_path) as file:
        file_data = file.readlines()[:56]
    for model_id in model_instance_names.split(';'):
        new_model_data = f'''\n\n# Target - {model_id}
# ----------------------
add_library({model_id} SHARED
    benchmark_model.c
)
target_include_directories({model_id}
    PRIVATE
        ${{DSE_CLIB_INCLUDE_DIR}}
        ../../../..
)
target_link_libraries({model_id}
    PRIVATE
        $<$<BOOL:${{WIN32}}>:${{modelc_link_lib}}>
)
install(TARGETS {model_id}
    LIBRARY DESTINATION
        ${{MODEL_PATH}}/lib
    COMPONENT
        {model_id}
)'''
        file_data.append(new_model_data)
    with open(cmake_file_path, "w") as stream:
        stream.writelines(file_data)


def update_model_file(model_instance_names):
    ''' Updates model.yaml file dynamically with new model instance data.'''
    docs = []
    for model in model_instance_names.split(';'):
        model_obj = {'kind': 'Model', 'metadata': {'name': model}, 'spec': {'runtime': {'dynlib': [{'os': 'linux', 'arch': 'amd64', 'path': 'lib/'+model+'.so'}, {'os': 'linux', 'arch': 'x86', 'path': 'lib/'+model+'.so'}, {'os': 'windows', 'arch': 'x64', 'path': 'bin/'+model+'.dll'}, {'os': 'windows', 'arch': 'x86', 'path': 'bin/'+model+'.dll'}]}, 'channels': [{'alias': 'model_channel', 'selectors': {'channel': 'test'}}]}}
        docs.append(model_obj)
    with open("_working/model.yaml", "w") as stream:
        yaml.dump_all(docs, stream)


def update_stack_file(model_instance_names, num_of_models):
    ''' function to update the stack.yaml file with dynamically generated model instance names data'''
    updated_docs = []
    with open("data/stack.yaml") as file:
        docs = yaml.load_all(file)
        model_instance_names = model_instance_names.split(';')
        for doc in docs:
            try:
                doc['spec']['models'][0]['channels'][0]['expectedModelCount'] = int(num_of_models)
                doc['spec']['models'] = doc['spec']['models'][:1]
                uid = 43
                for model_instance in model_instance_names:
                    instance = {
                        'name': model_instance, 'uid': uid, 'model': {'name': model_instance_names[0] if USE_SAME_MODEL_SOURCE_OBJECT else model_instance},
                        'channels': [{'name': 'test', 'alias': 'model_channel'}]}
                    uid += 1
                    doc['spec']['models'].append(instance)
            except Exception as _:
                pass
            updated_docs.append(doc)
    with open("_working/stack.yaml", "w") as stream:
        yaml.dump_all(updated_docs, stream)


def raise_exception(msg):
    raise Exception(msg)


def check_type(value, check_if):
    status = False
    if check_if == 'int':
        try :
            int(value)
        except ValueError:
            status = True
    if check_if == 'float':
        try :
            float(value)
        except ValueError:
            status = True
    return status


def validate_input(simbus_uri, num_of_models, step_size, end_time, signal_count, signal_change):
    simbus_uri = simbus_uri.strip()
    if simbus_uri not in ('redis://redis:6379', 'unix:///tmp/redis/redis.sock') :
        raise_exception('Invalid simbus URI.')
        return False
    if check_type(num_of_models, 'int'):
        raise_exception('num_of_models value should be integer')
        return False
    if check_type(step_size, 'float'):
        raise_exception('step_size value should be float/int')
        return False
    if check_type(end_time, 'float'):
        raise_exception('end_time value should be float/int')
        return False
    if check_type(signal_count, 'int'):
        raise_exception('signal_count value should be integer')
        return False
    if check_type(signal_change, 'int'):
        raise_exception('signal_change value should be integer')
        return False


def run_benchmark(input_data):
    prev_signal_count = 0
    prev_model_count = 0
    for line in input_data:
        data = line.split(',')
        if '' in data :
            continue # checks if any col value missing or any empty row in inpuit csv file
        simbus_uri = data[0]
        num_of_models = data[1]
        step_size = data[2]
        end_time = data[3]
        signal_count = data[4]
        signal_change = data[5]
        out_file = data[6]
        if validate_input(simbus_uri, num_of_models, step_size, end_time, signal_count, signal_change) is False :
            continue

        model_instance_names = generate_model_instance_names(num_of_models)
        update_stack_file(model_instance_names, num_of_models)
        update_cmakelist(model_instance_names)
        update_model_file(model_instance_names)
        # creates .env file to set the env variables for docker-compose.yaml
        with open('_working/.env', 'w') as file:
            content = [
                f'SIMBUS_URI={simbus_uri}\n',
                f'NUM_OF_MODELS={num_of_models}\n',
                f'STEP_SIZE={step_size}\n',
                f'END_TIME={end_time}\n',
                f'SIGNAL_COUNT={signal_count}\n',
                f'SIGNAL_CHANGE={signal_change}\n',
                f'MODEL_INSTANCE_NAMES={model_instance_names}\n',
                f'OUT_FILE={out_file}\n',
            ]
            file.writelines(content)

        # generates signal_group.yaml file only if there is a change in signal_count from the above row in input.csv
        if signal_count != prev_signal_count or num_of_models != prev_model_count:
            signal_gen.main(signal_count)
            os.chdir("../../..")
            os.system('make')
            os.chdir("tests/pytest/benchmark")
            prev_signal_count = signal_count
            prev_model_count = num_of_models
        os.system('make benchmark')


def process_data_for_plotting(df):
    signal_change_range = df['signal change'].unique()
    gp_df = df.groupby(['scenario', 'transport', 'models'])
    cols = ['scenario', 'transport', '']
    cols.extend([''] * len(signal_change_range))
    data_list = []
    prev_transport_name = ''
    prev_scenario_name = ''
    for gp_by, group in gp_df:
        row_items = [gp_by[0], gp_by[1], gp_by[2]]
        for row in group.iterrows():
            sim_steps_per_sec = row[1]['rate (steps/sec)']
            # signal_change = row[1]['signal change']
            row_items.extend([sim_steps_per_sec])
        if prev_transport_name != gp_by[1] or prev_scenario_name != gp_by[0]:
            tmp_header = ['', '', '']
            tmp_header.extend(signal_change_range)
            data_list.append(cols)
            data_list.append(tmp_header)
            prev_scenario_name = gp_by[0]
            prev_transport_name = gp_by[1]
        data_list.append(row_items)
    data_list.append(data_list[0]) # needed for splitting to different dfs
    data_list = data_list[1:]
    df_list = []
    tmp_list = []
    for row_list in data_list:
        if row_list[0] == 'scenario':
            tmp_df = pd.DataFrame(tmp_list, columns = cols)
            df_list.append(tmp_df)
            tmp_list.clear()
        else:
            tmp_list.append(row_list)
    return df_list


def plot_to_html(df):
    scenario = df['scenario'][1]
    transport = df['transport'][1]
    title = scenario.replace('_', ' ').upper() + ' - ' + transport.split('_')[-1].upper()
    data = df.drop(df.columns[[0, 1]], axis=1)
    x_data = list(data.iloc[0])[1:] # signal range
    y_data = list(data.iloc[:, 0])[1:] # Num. of models
    z_data = data.iloc[1: , 1:].values.tolist() # speed - sim.steps per sec.

    fig = go.Figure(go.Surface(

        x = x_data,
        y = y_data,
        z = z_data
    ))
    fig.update_layout(
            scene = {
                "xaxis": {"nticks": 20},
                "zaxis": {
                    "tickmode" : "array",
                    },
                "yaxis": {
                        "tickmode" : "array",
                        },
                'camera_eye': {"x": .5, "y": -1.5, "z": 0.5},
                "aspectratio": {"x": 1, "y": 1, "z": 0.2},
                "xaxis_title": 'Signal Change',
                "zaxis_title": 'Speed(sim. steps per sec.)',
                "yaxis_title": 'Num. of Models',
            },
            margin=dict(l=50, r=50, b=50, t=50, pad=0),
            autosize=False, width=900, height=500,
            title=title, title_x=0.5
            )
    return fig


def plot_to_excel(plot_data, csv_data, max_df_val, file_name):
    scenario = plot_data['scenario'][1]
    transport = plot_data['transport'][1]
    plot_data = plot_data.drop(columns = ['scenario', 'transport'])
    row_count, col_count = plot_data.shape
    out_file_name = f'_out/{file_name}.xlsx'
    sheet_name = '_'.join(scenario.split('_')[1:3]) + '_' + transport.split('_')[-1]
    if os.path.exists(out_file_name) is False :
        wb = Workbook()
        ws = wb.active
        for row in dataframe_to_rows(csv_data, index=False, header=True):# loading the complete plot data in first sheet.
            ws.append(row)
        ws.title = 'Plot_Data'
        wb.save(out_file_name)
    wb = load_workbook(out_file_name)
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    data = plot_data.values.tolist()
    for row in data:
        ws.append(row)
    chart = SurfaceChart3D()
    labels = Reference(ws, min_col = 1, min_row = 2, max_row = row_count)
    data = Reference(ws, min_col = 2, max_col = col_count, min_row = 1, max_row = row_count)
    chart.add_data(data, titles_from_data = True)
    chart.set_categories(labels)
    chart.title = (scenario.replace('_', ' ') + ' - ' + transport.replace('_', ' ')).title()
    chart.style = 26
    chart.height = 10 # default is 7.5
    chart.width = 20 # default is 15
    chart.y_axis.title = "Sim. Steps per sec."
    chart.x_axis.title = "Num. Of Models"
    chart.z_axis.title = "Signal Change"

    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = max_df_val
    # chart.y_axis.scaling.unit = 100
    if os.path.exists(f"_out/{file_name}.csv"):
        os.remove(f"_out/{file_name}.csv")
    ws.add_chart(chart, "K2")
    wb.active = wb[sheet_name]
    wb.save(out_file_name)


if __name__ == '__main__':
    out_type = 'excel' # 'excel' or 'html' (if 'html' then plotly lib is used for plotting else openpyxl)
    input_data = read_input_file()
    run_benchmark(input_data)
    print('Plotting Surface charts...')
    out_files = [file for file in glob.glob("_out/*.csv")] # input files

    def gen_plot_file(fig, file_path):
        '''Generates the ploted output file'''
        file_path = f"{file_path}.html"
        with open(file_path, 'a') as f :
            f.write(fig.to_html(full_html=False, include_plotlyjs='cdn'))

    def roundup(x):
        '''rounding the max y-axis value to its closest hundreds'''
        return int(math.ceil(x / 100.0)) * 100

    def get_max_df_value(df_list):
        '''To fetch the max value from dataframe to sync y-axis values across plots'''
        max_df_val = 0
        for df in df_list:
            df_max = df.iloc[1:].drop(df.columns[[0, 1]], axis=1).to_numpy().max()
            max_df_val = df_max if df_max > max_df_val else max_df_val
        max_df_val = roundup(max_df_val)
        return max_df_val

    for file in out_files:
        file_name = file.split('/')[-1].rstrip('.csv')
        file_path = f'_out/plot_{file_name}'
        if out_type == 'html' :
            if os.path.exists(f"{file_path}.html") :
                os.remove(f"{file_path}.html") # deletes the file is exists, else the same file will get appened in each alternate try.
        csv_data = pd.read_csv(file)
        df_list = process_data_for_plotting(csv_data)
        max_df_val = get_max_df_value(df_list)
        for df in df_list:
            if out_type == 'excel' :
                plot_to_excel(df, csv_data, max_df_val, file_name)
            else :
                fig = plot_to_html(df)
                gen_plot_file(fig, file_path)
    print(f'Plotted Surface charts.')
